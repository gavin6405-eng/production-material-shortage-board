import io
import re
from copy import copy
from datetime import date, datetime

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="超慧科技｜機台容量管理", page_icon="🏭", layout="wide")

st.markdown(
    """
    <style>
    .main-title {font-size: 2rem; font-weight: 800; margin-bottom: .2rem;}
    .sub-title {color: #64748b; margin-bottom: 1rem;}
    div[data-testid="stMetric"] {background:#f8fafc; border:1px solid #e2e8f0; padding:14px; border-radius:12px;}
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown('<div class="main-title">🏭 超慧科技機台容量管理</div>', unsafe_allow_html=True)
st.markdown(
    '<div class="sub-title">製令發料日＝進機（＋計算台數）；出機日＝出機（－計算台數），自動計算每日結存與目前實際台數。</div>',
    unsafe_allow_html=True,
)

ALIASES = {
    "發料日期": ["發料日期", "製令發料日", "發料日", "預計發料日"],
    "製令": ["製令", "製令號", "工單", "工單號"],
    "計算台數": ["計算台數", "台數", "機台數", "數量"],
    "出機日": ["出機日", "出貨日", "預計出機日", "預計出貨日"],
    "實際台數": ["實際台數", "在廠台數", "累計台數"],
    "容量上限": ["容量上限", "容量上限（台）", "最大容量"],
}


def norm_text(value):
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value)).strip().lower()


def find_header_row(ws, max_scan=40):
    required = {norm_text(x) for x in ALIASES["發料日期"]}
    order_names = {norm_text(x) for x in ALIASES["製令"]}
    qty_names = {norm_text(x) for x in ALIASES["計算台數"]}
    for r in range(1, min(ws.max_row, max_scan) + 1):
        vals = {norm_text(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)}
        if vals & required and vals & order_names and vals & qty_names:
            return r
    return None


def map_columns(ws, header_row):
    headers = {c: norm_text(ws.cell(header_row, c).value) for c in range(1, ws.max_column + 1)}
    result = {}
    for canonical, aliases in ALIASES.items():
        alias_set = {norm_text(x) for x in aliases}
        for c, header in headers.items():
            if header in alias_set:
                result[canonical] = c
                break
    return result


def parse_date(value):
    if value is None or value == "":
        return pd.NaT
    if isinstance(value, (datetime, date)):
        return pd.Timestamp(value).normalize()
    text = str(value).strip()
    if text.upper() in {"TBD", "N/A", "NA", "待定", "未定", "-", "--"}:
        return pd.NaT
    return pd.to_datetime(value, errors="coerce").normalize()


def parse_qty(value):
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        text = re.sub(r"[^0-9.\-]", "", str(value))
        try:
            return float(text) if text else 0.0
        except ValueError:
            return 0.0


def clean_number(value):
    return int(value) if float(value).is_integer() else round(float(value), 2)


def detect_capacity(ws):
    for r in range(1, min(ws.max_row, 30) + 1):
        for c in range(1, min(ws.max_column, 10) + 1):
            if norm_text(ws.cell(r, c).value) in {norm_text(x) for x in ALIASES["容量上限"]}:
                candidate = ws.cell(r, c + 1).value
                try:
                    return float(candidate)
                except (TypeError, ValueError):
                    pass
    try:
        return float(ws["B5"].value)
    except (TypeError, ValueError):
        return 82.0


def read_source(file_bytes):
    wb = load_workbook(io.BytesIO(file_bytes), data_only=False)
    candidates = []
    for ws in wb.worksheets:
        header_row = find_header_row(ws)
        if header_row:
            candidates.append((ws.title, header_row, map_columns(ws, header_row)))
    if not candidates:
        raise ValueError("找不到包含『發料日期／製令／計算台數』的資料表。")
    return wb, candidates


def extract_records(ws, header_row, cols):
    records = []
    for r in range(header_row + 1, ws.max_row + 1):
        release_raw = ws.cell(r, cols["發料日期"]).value
        order_raw = ws.cell(r, cols["製令"]).value
        qty_raw = ws.cell(r, cols["計算台數"]).value
        exit_raw = ws.cell(r, cols.get("出機日", 0)).value if cols.get("出機日") else None

        if all(v in (None, "") for v in (release_raw, order_raw, qty_raw, exit_raw)):
            continue

        release_date = parse_date(release_raw)
        exit_date = parse_date(exit_raw)
        qty = parse_qty(qty_raw)
        order_no = "" if order_raw is None else str(order_raw).strip()

        records.append(
            {
                "原始列號": r,
                "製令": order_no,
                "發料日期": release_date,
                "出機日": exit_date,
                "計算台數": qty,
                "原發料值": release_raw,
                "原出機值": exit_raw,
            }
        )
    return pd.DataFrame(records)


def build_event_tables(records, initial_qty=0.0):
    events = []
    errors = []
    for _, row in records.iterrows():
        order_no = row["製令"]
        qty = float(row["計算台數"])
        release_dt = row["發料日期"]
        exit_dt = row["出機日"]
        source_row = int(row["原始列號"])

        if qty <= 0:
            errors.append({"原始列號": source_row, "製令": order_no, "異常": "計算台數空白或小於等於 0"})
            continue
        if pd.isna(release_dt):
            errors.append({"原始列號": source_row, "製令": order_no, "異常": "發料日期無法辨識"})
        else:
            events.append(
                {
                    "日期": release_dt,
                    "事件順序": 2,
                    "事件": "進機",
                    "製令": order_no,
                    "進機台數": qty,
                    "出機台數": 0.0,
                    "異動台數": qty,
                    "原始列號": source_row,
                }
            )
        if not pd.isna(exit_dt):
            if not pd.isna(release_dt) and exit_dt < release_dt:
                errors.append({"原始列號": source_row, "製令": order_no, "異常": "出機日早於發料日期"})
            events.append(
                {
                    "日期": exit_dt,
                    "事件順序": 1,
                    "事件": "出機",
                    "製令": order_no,
                    "進機台數": 0.0,
                    "出機台數": qty,
                    "異動台數": -qty,
                    "原始列號": source_row,
                }
            )

    event_df = pd.DataFrame(events)
    if event_df.empty:
        event_df = pd.DataFrame(columns=["日期", "事件順序", "事件", "製令", "進機台數", "出機台數", "異動台數", "原始列號", "事件後總量"])
        daily_df = pd.DataFrame(columns=["日期", "進機台數", "出機台數", "當日淨增減", "每日結存"])
        return event_df, daily_df, pd.DataFrame(errors)

    # 同一天先扣出機、再加進機；每日最終結存不受先後順序影響。
    event_df = event_df.sort_values(["日期", "事件順序", "原始列號", "製令"]).reset_index(drop=True)
    event_df["事件後總量"] = initial_qty + event_df["異動台數"].cumsum()

    daily_df = (
        event_df.groupby("日期", as_index=False)
        .agg(進機台數=("進機台數", "sum"), 出機台數=("出機台數", "sum"), 當日淨增減=("異動台數", "sum"))
        .sort_values("日期")
    )
    daily_df["每日結存"] = initial_qty + daily_df["當日淨增減"].cumsum()
    return event_df, daily_df, pd.DataFrame(errors)


def status_text(qty, capacity):
    if capacity <= 0:
        return "未設定"
    rate = qty / capacity
    if qty > capacity:
        return "超載"
    if rate >= 0.96:
        return "紅燈"
    if rate >= 0.91:
        return "橘燈"
    if rate >= 0.81:
        return "黃燈"
    return "綠燈"


def style_output_sheet(ws, widths):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def copy_style(source, target):
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)


def export_workbook(original_bytes, sheet_name, header_row, cols, records, event_df, daily_df, errors_df, capacity, initial_qty, as_of):
    wb = load_workbook(io.BytesIO(original_bytes), data_only=False)
    ws = wb[sheet_name]

    # 確保「實際台數」欄存在；沒有時新增。
    actual_col = cols.get("實際台數")
    if not actual_col:
        actual_col = ws.max_column + 1
        ws.cell(header_row, actual_col, "實際台數")
        if actual_col > 1:
            copy_style(ws.cell(header_row, actual_col - 1), ws.cell(header_row, actual_col))

    # 新增「目前是否在廠」與「截至查詢日台數影響」。
    in_factory_col = ws.max_column + 1
    impact_col = ws.max_column + 2
    ws.cell(header_row, in_factory_col, "目前是否在廠")
    ws.cell(header_row, impact_col, f"截至{as_of:%Y-%m-%d}台數影響")
    copy_style(ws.cell(header_row, actual_col), ws.cell(header_row, in_factory_col))
    copy_style(ws.cell(header_row, actual_col), ws.cell(header_row, impact_col))

    daily_balance = daily_df.set_index("日期")["每日結存"].to_dict() if not daily_df.empty else {}

    # 寫回每筆製令在「發料日當日結束後」的全廠結存。
    for _, rec in records.iterrows():
        r = int(rec["原始列號"])
        release_dt = rec["發料日期"]
        exit_dt = rec["出機日"]
        qty = float(rec["計算台數"])
        if not pd.isna(release_dt):
            ws.cell(r, actual_col, clean_number(daily_balance.get(release_dt, initial_qty)))
        else:
            ws.cell(r, actual_col, None)

        is_active = (not pd.isna(release_dt)) and (release_dt.date() <= as_of) and (pd.isna(exit_dt) or exit_dt.date() > as_of)
        ws.cell(r, in_factory_col, "是" if is_active else "否")
        ws.cell(r, impact_col, clean_number(qty) if is_active else 0)

        if r > header_row + 1:
            copy_style(ws.cell(r - 1, actual_col), ws.cell(r, actual_col))
            copy_style(ws.cell(r - 1, in_factory_col), ws.cell(r, in_factory_col))
            copy_style(ws.cell(r - 1, impact_col), ws.cell(r, impact_col))

    ws.column_dimensions[get_column_letter(in_factory_col)].width = 15
    ws.column_dimensions[get_column_letter(impact_col)].width = 23

    # 更新上方摘要值。
    current_qty = initial_qty
    if not daily_df.empty:
        eligible = daily_df[daily_df["日期"].dt.date <= as_of]
        if not eligible.empty:
            current_qty = float(eligible.iloc[-1]["每日結存"])
    ws["B6"] = clean_number(current_qty)
    ws["B7"] = "=IFERROR(B6/B5,0)"
    ws["B8"] = '=IF(B6>B5,"超載",IF(B7>=0.96,"接近或已滿載",IF(B7>=0.91,"空間高負載",IF(B7>=0.81,"提前規劃移機或入庫","正常運作"))))'

    # 重建分析工作表。
    for name in ["每日容量彙整", "進出事件明細", "資料異常檢查"]:
        if name in wb.sheetnames:
            del wb[name]

    daily_ws = wb.create_sheet("每日容量彙整")
    daily_headers = ["日期", "進機台數", "出機台數", "當日淨增減", "每日結存", "容量上限", "使用率", "管理燈號"]
    daily_ws.append(daily_headers)
    for _, row in daily_df.iterrows():
        balance = float(row["每日結存"])
        daily_ws.append([
            row["日期"].to_pydatetime(), clean_number(row["進機台數"]), clean_number(row["出機台數"]),
            clean_number(row["當日淨增減"]), clean_number(balance), clean_number(capacity),
            balance / capacity if capacity else None, status_text(balance, capacity),
        ])
    style_output_sheet(daily_ws, [13, 12, 12, 14, 12, 12, 12, 13])
    for c in daily_ws["A"][1:]:
        c.number_format = "yyyy-mm-dd"
    for c in daily_ws["G"][1:]:
        c.number_format = "0.0%"

    event_ws = wb.create_sheet("進出事件明細")
    event_headers = ["日期", "事件", "製令", "進機台數", "出機台數", "異動台數", "事件後總量", "原始列號"]
    event_ws.append(event_headers)
    for _, row in event_df.iterrows():
        event_ws.append([
            row["日期"].to_pydatetime(), row["事件"], row["製令"], clean_number(row["進機台數"]),
            clean_number(row["出機台數"]), clean_number(row["異動台數"]), clean_number(row["事件後總量"]), int(row["原始列號"]),
        ])
    style_output_sheet(event_ws, [13, 10, 20, 12, 12, 12, 14, 12])
    for c in event_ws["A"][1:]:
        c.number_format = "yyyy-mm-dd"

    if not errors_df.empty:
        err_ws = wb.create_sheet("資料異常檢查")
        err_ws.append(list(errors_df.columns))
        for row in errors_df.itertuples(index=False, name=None):
            err_ws.append(list(row))
        style_output_sheet(err_ws, [12, 20, 34])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue(), current_qty


uploaded = st.file_uploader("上傳機台容量管理 Excel", type=["xlsx", "xlsm"])

with st.expander("計算規則", expanded=False):
    st.markdown(
        """
- **製令發料日**：當天進機，台數以正數加入。
- **出機日／出貨日**：當天出機，同一筆製令的計算台數以負數扣除。
- **每日結存**：前一日結存＋當日進機－當日出機。
- 原檔「實際台數」會更新為該筆製令**發料日當日結束後的全廠結存**。
- 另新增「每日容量彙整」與「進出事件明細」，可完整查看日期交錯後的真正累計數。
- 出機日為 `TBD`、空白或無法辨識時，視為尚未出機。
        """
    )

if uploaded is None:
    st.info("請先上傳 Excel 檔案。")
    st.stop()

file_bytes = uploaded.getvalue()

try:
    wb_preview, candidates = read_source(file_bytes)
except Exception as exc:
    st.error(f"讀取失敗：{exc}")
    st.stop()

candidate_labels = [f"{name}（標題列：{header_row}）" for name, header_row, _ in candidates]
selected_label = st.selectbox("選擇資料工作表", candidate_labels, index=0)
selected_idx = candidate_labels.index(selected_label)
sheet_name, header_row, cols = candidates[selected_idx]
ws_preview = wb_preview[sheet_name]
capacity_default = detect_capacity(ws_preview)

c1, c2, c3 = st.columns(3)
with c1:
    capacity = st.number_input("容量上限（台）", min_value=0.0, value=float(capacity_default), step=1.0)
with c2:
    initial_qty = st.number_input("期初在廠台數", min_value=0.0, value=0.0, step=1.0, help="只填寫未包含在明細內、但計算起始日前已存在的機台。")
with c3:
    as_of = st.date_input("目前台數計算截止日", value=date.today())

try:
    records = extract_records(ws_preview, header_row, cols)
    event_df, daily_df, errors_df = build_event_tables(records, initial_qty=initial_qty)
except Exception as exc:
    st.error(f"資料處理失敗：{exc}")
    st.stop()

eligible = daily_df[daily_df["日期"].dt.date <= as_of] if not daily_df.empty else daily_df
current_qty = float(eligible.iloc[-1]["每日結存"]) if not eligible.empty else float(initial_qty)
rate = current_qty / capacity if capacity else 0
active_orders = records[
    records["發料日期"].notna()
    & (records["發料日期"].dt.date <= as_of)
    & (records["出機日"].isna() | (records["出機日"].dt.date > as_of))
]

m1, m2, m3, m4 = st.columns(4)
m1.metric("目前實際台數", clean_number(current_qty))
m2.metric("容量上限", clean_number(capacity))
m3.metric("容量使用率", f"{rate:.1%}")
m4.metric("目前狀態", status_text(current_qty, capacity))

if current_qty < 0:
    st.warning("目前結存出現負數，請檢查是否有出機日早於發料日、重複出機或期初台數未填。")
if current_qty > capacity and capacity > 0:
    st.error(f"目前超出容量 {clean_number(current_qty - capacity)} 台。")

left, right = st.columns([1.25, 1])
with left:
    st.subheader("每日容量趨勢")
    if daily_df.empty:
        st.info("沒有可計算的日期資料。")
    else:
        chart_df = daily_df.set_index("日期")[["每日結存"]]
        st.line_chart(chart_df, use_container_width=True)
with right:
    st.subheader("目前在廠製令")
    show_active = active_orders[["製令", "發料日期", "出機日", "計算台數"]].copy()
    if show_active.empty:
        st.info("截止日沒有在廠製令。")
    else:
        show_active["發料日期"] = show_active["發料日期"].dt.strftime("%Y-%m-%d")
        show_active["出機日"] = show_active["出機日"].dt.strftime("%Y-%m-%d").fillna("TBD")
        st.dataframe(show_active, use_container_width=True, hide_index=True)

with st.expander("查看每日進出彙整", expanded=True):
    display_daily = daily_df.copy()
    if not display_daily.empty:
        display_daily["日期"] = display_daily["日期"].dt.strftime("%Y-%m-%d")
        display_daily["容量上限"] = capacity
        display_daily["使用率"] = display_daily["每日結存"] / capacity if capacity else 0
        display_daily["管理燈號"] = display_daily["每日結存"].apply(lambda x: status_text(float(x), capacity))
        for col in ["進機台數", "出機台數", "當日淨增減", "每日結存"]:
            display_daily[col] = display_daily[col].map(clean_number)
        st.dataframe(display_daily, use_container_width=True, hide_index=True)

with st.expander("查看進出事件明細"):
    display_event = event_df.drop(columns=["事件順序"], errors="ignore").copy()
    if not display_event.empty:
        display_event["日期"] = display_event["日期"].dt.strftime("%Y-%m-%d")
        for col in ["進機台數", "出機台數", "異動台數", "事件後總量"]:
            display_event[col] = display_event[col].map(clean_number)
    st.dataframe(display_event, use_container_width=True, hide_index=True)

if not errors_df.empty:
    with st.expander(f"⚠️ 資料異常檢查（{len(errors_df)} 筆）", expanded=True):
        st.dataframe(errors_df, use_container_width=True, hide_index=True)

try:
    output_bytes, exported_current_qty = export_workbook(
        original_bytes=file_bytes,
        sheet_name=sheet_name,
        header_row=header_row,
        cols=cols,
        records=records,
        event_df=event_df,
        daily_df=daily_df,
        errors_df=errors_df,
        capacity=capacity,
        initial_qty=initial_qty,
        as_of=as_of,
    )
    output_name = f"機台容量管理_進出計算_{datetime.now():%Y%m%d_%H%M}.xlsx"
    st.download_button(
        "⬇️ 下載已完成計算的 Excel",
        data=output_bytes,
        file_name=output_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )
except Exception as exc:
    st.error(f"輸出 Excel 失敗：{exc}")
